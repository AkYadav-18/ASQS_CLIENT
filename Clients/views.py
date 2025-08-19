from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import Client
from Partners.models import Partner
from Standards.models import Standards
import re
import json
from datetime import datetime
from django.contrib.auth.decorators import login_required
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Create your views here.

def validate_date_format(date_string):
    """Validate date string in yyyy-mm-dd format"""
    if not date_string:
        return False

    # Check format with regex
    pattern = r'^\d{4}-\d{2}-\d{2}$'
    if not re.match(pattern, date_string):
        return False

    # Check if it's a valid date
    try:
        datetime.strptime(date_string, '%Y-%m-%d')
        return True
    except ValueError:
        return False


@login_required
def add_client(request, template_name='Clients/add_client.html'):
    # Get all partners and standards for dropdowns
    partners = Partner.objects.all()
    standards = Standards.objects.all()

    # Get all clients and apply filters
    clients = Client.objects.all()

    # Apply filters if provided
    status_filter = request.GET.get('status_filter')
    country_filter = request.GET.get('country_filter')

    if status_filter:
        clients = clients.filter(certification_status=status_filter)

    if country_filter:
        clients = clients.filter(country=country_filter)

    clients = clients.order_by('-created_at')

    # Add pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(clients, 20)  # Show 10 clients per page

    try:
        clients = paginator.page(page)
    except PageNotAnInteger:
        clients = paginator.page(1)
    except EmptyPage:
        clients = paginator.page(paginator.num_pages)

    # Check if this is an edit request
    edit_client_id = request.GET.get('edit')
    edit_client = None
    if edit_client_id:
        try:
            edit_client = Client.objects.get(id=edit_client_id)
        except Client.DoesNotExist:
            messages.error(request, 'Client not found.')
            return redirect('clients')

    # Define country choices (common countries for certification)
    COUNTRY_CHOICES = [
        ('', 'Select Country'),
        ('US', 'United States'),
        ('CA', 'Canada'),
        ('UK', 'United Kingdom'),
        ('DE', 'Germany'),
        ('FR', 'France'),
        ('IT', 'Italy'),
        ('ES', 'Spain'),
        ('NL', 'Netherlands'),
        ('BE', 'Belgium'),
        ('CH', 'Switzerland'),
        ('AT', 'Austria'),
        ('SE', 'Sweden'),
        ('NO', 'Norway'),
        ('DK', 'Denmark'),
        ('FI', 'Finland'),
        ('AU', 'Australia'),
        ('NZ', 'New Zealand'),
        ('JP', 'Japan'),
        ('KR', 'South Korea'),
        ('CN', 'China'),
        ('IN', 'India'),
        ('SG', 'Singapore'),
        ('MY', 'Malaysia'),
        ('TH', 'Thailand'),
        ('BR', 'Brazil'),
        ('MX', 'Mexico'),
        ('AR', 'Argentina'),
        ('ZA', 'South Africa'),
        ('EG', 'Egypt'),
        ('AE', 'United Arab Emirates'),
    ]

    if request.method == 'POST':
        try:
            # Check if this is an edit or add operation
            edit_id = request.POST.get('edit_client_id')
            if edit_id:
                try:
                    client = Client.objects.get(id=edit_id)
                    operation = 'updated'
                except Client.DoesNotExist:
                    messages.error(request, 'Client not found for editing.')
                    return redirect('clients')
            else:
                # Create new client instance
                client = Client()
                operation = 'added'

            # Basic Info
            client.name = request.POST.get('name', '').strip()
            client.email = request.POST.get('email', '').strip()
            client.phone = request.POST.get('phone', '').strip()
            client.address = request.POST.get('address', '').strip()
            client.address2 = request.POST.get('address2', '').strip()
            client.address3 = request.POST.get('address3', '').strip()

            # Certification Details
            client.certification_number = request.POST.get('certification_number', '').strip()
            client.country = request.POST.get('country', '').strip()
            client.standard = request.POST.get('standard', '').strip()

            # Handle dates (yyyy-mm-dd format)
            cert_date = request.POST.get('certification_date', '').strip()
            if cert_date:
                if validate_date_format(cert_date):
                    client.certification_date = cert_date
                else:
                    messages.error(request, 'Invalid certification date format. Please use yyyy-mm-dd.')
                    return render(request, template_name, {
                        'partners': partners,
                        'standards': standards,
                        'clients': clients,
                        'countries': COUNTRY_CHOICES,
                        'form_data': request.POST
                    })

            expiry_date = request.POST.get('expiry_date', '').strip()
            if expiry_date:
                if validate_date_format(expiry_date):
                    client.expiry_date = expiry_date
                else:
                    messages.error(request, 'Invalid expiry date format. Please use yyyy-mm-dd.')
                    return render(request, template_name, {
                        'partners': partners,
                        'standards': standards,
                        'clients': clients,
                        'countries': COUNTRY_CHOICES,
                        'form_data': request.POST
                    })

            recert_date = request.POST.get('recertification_date', '').strip()
            if recert_date:
                if validate_date_format(recert_date):
                    client.recertification_date = recert_date
                else:
                    messages.error(request, 'Invalid recertification date format. Please use yyyy-mm-dd.')
                    return render(request, template_name, {
                        'partners': partners,
                        'standards': standards,
                        'clients': clients,
                        'countries': COUNTRY_CHOICES,
                        'form_data': request.POST
                    })

            issue_date = request.POST.get('issue_date', '').strip()
            if issue_date:
                if validate_date_format(issue_date):
                    client.issue_date = issue_date
                else:
                    messages.error(request, 'Invalid issue date format. Please use yyyy-mm-dd.')
                    return render(request, template_name, {
                        'partners': partners,
                        'standards': standards,
                        'clients': clients,
                        'countries': COUNTRY_CHOICES,
                        'form_data': request.POST
                    })

            # Other fields
            client.accreditation_body = request.POST.get('accreditation_body', '').strip()
            client.scope = request.POST.get('scope', '').strip()
            client.certification_status = request.POST.get('certification_status', 'active')
            client.certification_type = request.POST.get('certification_type', '')

            # Partner
            partner_id = request.POST.get('partner')
            if partner_id:
                try:
                    client.partner = Partner.objects.get(id=partner_id)
                except Partner.DoesNotExist:
                    pass

            # Set created_by if user is authenticated
            if request.user.is_authenticated:
                client.created_by = request.user
                # Track who made the change for status history
                client._changed_by = request.user

            # Validate required fields
            if not client.name:
                messages.error(request, 'Name is required.')
                return render(request, template_name, {
                    'partners': partners,
                    'standards': standards,
                    'clients': clients,
                    'countries': COUNTRY_CHOICES,
                    'form_data': request.POST
                })

            if not client.certification_number:
                messages.error(request, 'Certification number is required.')
                return render(request, template_name, {
                    'partners': partners,
                    'standards': standards,
                    'clients': clients,
                    'countries': COUNTRY_CHOICES,
                    'form_data': request.POST
                })

            # Save the client
            client.save()
            messages.success(request, f'Client "{client.name}" has been {operation} successfully!')
            return redirect('clients')

        except Exception as e:
            messages.error(request, f'Error adding client: {str(e)}')

    return render(request, template_name, {
        'partners': partners,
        'standards': standards,
        'clients': clients,
        'countries': COUNTRY_CHOICES,
        'edit_client': edit_client,
    })


@login_required
def get_client(request, client_id):
    """Get client data for editing"""
    try:
        client = Client.objects.get(id=client_id)
        client_data = {
            'id': client.id,
            'name': client.name,
            'email': client.email,
            'phone': client.phone,
            'address': client.address,
            'address2': client.address2,
            'address3': client.address3,
            'certification_number': client.certification_number,
            'country': client.country,
            'standard': client.standard,
            'certification_date': client.certification_date.strftime('%Y-%m-%d') if client.certification_date else '',
            'expiry_date': client.expiry_date.strftime('%Y-%m-%d') if client.expiry_date else '',
            'recertification_date': client.recertification_date.strftime('%Y-%m-%d') if client.recertification_date else '',
            'issue_date': client.issue_date.strftime('%Y-%m-%d') if client.issue_date else '',
            'accreditation_body': client.accreditation_body,
            'scope': client.scope,
            'certification_status': client.certification_status,
            'certification_type': client.certification_type,
            'partner_id': client.partner.id if client.partner else '',
        }
        return JsonResponse({'success': True, 'client': client_data})
    except Client.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Client not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def delete_client(request, client_id):
    """Delete client"""
    if request.method == 'POST':
        try:
            client = Client.objects.get(id=client_id)
            client_name = client.name
            client.delete()
            return JsonResponse({'success': True, 'message': f'Client "{client_name}" deleted successfully'})
        except Client.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Client not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def export_clients_excel(request):
    """Export clients to Excel file"""
    try:
        # Get all clients with applied filters
        clients = Client.objects.all()

        # Apply filters if provided
        status_filter = request.GET.get('status_filter')
        country_filter = request.GET.get('country_filter')

        if status_filter:
            clients = clients.filter(certification_status=status_filter)

        if country_filter:
            clients = clients.filter(country=country_filter)

        clients = clients.order_by('-created_at')

        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Clients"

        # Define headers
        headers = [
            'Name', 'Email', 'Phone', 'Address', 'Address 2', 'Address 3',
            'Certification Number', 'Country', 'Standard', 'Status',
            'Certification Date', 'Expiry Date', 'Recertification Date',
            'Issue Date', 'Accreditation Body', 'Scope', 'Certification Type',
            'Partner', 'Created At'
        ]

        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add data rows
        for row_num, client in enumerate(clients, 2):
            data = [
                client.name,
                client.email or '',
                client.phone or '',
                client.address,
                client.address2 or '',
                client.address3 or '',
                client.certification_number,
                client.country,
                client.standard or '',
                client.get_certification_status_display(),
                client.certification_date.strftime('%Y-%m-%d') if client.certification_date else '',
                client.expiry_date.strftime('%Y-%m-%d') if client.expiry_date else '',
                client.recertification_date.strftime('%Y-%m-%d') if client.recertification_date else '',
                client.issue_date.strftime('%Y-%m-%d') if client.issue_date else '',
                client.accreditation_body or '',
                client.scope or '',
                client.get_certification_type_display() if client.certification_type else '',
                client.partner.partner_name if client.partner else '',
                client.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ]

            for col_num, value in enumerate(data, 1):
                worksheet.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="clients_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        # Save workbook to response
        workbook.save(response)
        return response

    except Exception as e:
        messages.error(request, f'Error exporting clients: {str(e)}')
        return redirect('clients')
