from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import Partner
import json
from django.contrib.auth.decorators import login_required
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create your views here.

@login_required
def add_partner(request, template_name='Partners/add_partner.html'):
    # Get all partners and apply filters
    partners = Partner.objects.all()

    # Apply filters if provided
    country_filter = request.GET.get('country_filter')

    if country_filter:
        partners = partners.filter(country__icontains=country_filter)

    partners = partners.order_by('-created_at')

    # Add pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(partners, 10)  # Show 10 partners per page

    try:
        partners = paginator.page(page)
    except PageNotAnInteger:
        partners = paginator.page(1)
    except EmptyPage:
        partners = paginator.page(paginator.num_pages)

    # Check if this is an edit request
    edit_partner_id = request.GET.get('edit')
    edit_partner = None
    if edit_partner_id:
        try:
            edit_partner = Partner.objects.get(id=edit_partner_id)
        except Partner.DoesNotExist:
            messages.error(request, 'Partner not found.')
            return redirect('partners')

    # Define country choices (common countries)
    COUNTRY_CHOICES = [
        ('', 'Select Country'),
        ('US', 'United States'),
        ('CA', 'Canada'),
        ('GB', 'United Kingdom'),
        ('DE', 'Germany'),
        ('FR', 'France'),
        ('IT', 'Italy'),
        ('ES', 'Spain'),
        ('NL', 'Netherlands'),
        ('BE', 'Belgium'),
        ('CH', 'Switzerland'),
        ('AT', 'Austria'),
        ('SE', 'Sweden'),
        ('NO', 'Norway'),
        ('DK', 'Denmark'),
        ('FI', 'Finland'),
        ('AU', 'Australia'),
        ('NZ', 'New Zealand'),
        ('JP', 'Japan'),
        ('KR', 'South Korea'),
        ('CN', 'China'),
        ('IN', 'India'),
        ('SG', 'Singapore'),
        ('MY', 'Malaysia'),
        ('TH', 'Thailand'),
        ('BR', 'Brazil'),
        ('MX', 'Mexico'),
        ('AR', 'Argentina'),
        ('ZA', 'South Africa'),
        ('EG', 'Egypt'),
        ('AE', 'United Arab Emirates'),
    ]

    if request.method == 'POST':
        try:
            # Check if this is an edit or add operation
            edit_id = request.POST.get('edit_partner_id')
            if edit_id:
                try:
                    partner = Partner.objects.get(id=edit_id)
                    operation = 'updated'
                except Partner.DoesNotExist:
                    messages.error(request, 'Partner not found for editing.')
                    return redirect('partners')
            else:
                # Create new partner instance
                partner = Partner()
                operation = 'added'

            # Set form data
            partner.partner_name = request.POST.get('partner_name', '').strip()
            partner.company_name = request.POST.get('company_name', '').strip()
            partner.person_name = request.POST.get('person_name', '').strip()
            partner.email = request.POST.get('email', '').strip()
            partner.phone = request.POST.get('phone', '').strip()
            partner.address = request.POST.get('address', '').strip()
            partner.country = request.POST.get('country', '').strip()

            # Set created_by if user is authenticated
            if request.user.is_authenticated and not edit_id:
                partner.created_by = request.user

            # Validate required fields
            if not partner.partner_name:
                messages.error(request, 'Partner name is required.')
                return render(request, template_name, {
                    'partners': partners,
                    'countries': COUNTRY_CHOICES,
                    'form_data': request.POST
                })

            if not partner.company_name:
                messages.error(request, 'Company name is required.')
                return render(request, template_name, {
                    'partners': partners,
                    'countries': COUNTRY_CHOICES,
                    'form_data': request.POST
                })

            if not partner.email:
                messages.error(request, 'Email is required.')
                return render(request, template_name, {
                    'partners': partners,
                    'countries': COUNTRY_CHOICES,
                    'form_data': request.POST
                })

            # Save the partner
            partner.save()
            messages.success(request, f'Partner "{partner.partner_name}" has been {operation} successfully!')
            return redirect('partners')

        except Exception as e:
            messages.error(request, f'Error adding partner: {str(e)}')

    return render(request, template_name, {
        'partners': partners,
        'countries': COUNTRY_CHOICES,
        'edit_partner': edit_partner,
    })


@login_required
def get_partner(request, partner_id):
    """Get partner data for editing"""
    try:
        partner = Partner.objects.get(id=partner_id)
        partner_data = {
            'id': partner.id,
            'partner_name': partner.partner_name,
            'company_name': partner.company_name,
            'person_name': partner.person_name,
            'email': partner.email,
            'phone': partner.phone,
            'address': partner.address,
            'country': partner.country,
        }
        return JsonResponse({'success': True, 'partner': partner_data})
    except Partner.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Partner not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def delete_partner(request, partner_id):
    """Delete partner"""
    try:
        partner = Partner.objects.get(id=partner_id)
        partner_name = partner.partner_name
        partner.delete()
        return JsonResponse({'success': True, 'message': f'Partner "{partner_name}" has been deleted successfully!'})
    except Partner.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Partner not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def export_partners_excel(request):
    """Export partners to Excel file"""
    try:
        # Get all partners with applied filters
        partners = Partner.objects.all()

        # Apply filters if provided
        country_filter = request.GET.get('country_filter')

        if country_filter:
            partners = partners.filter(country__icontains=country_filter)

        partners = partners.order_by('-created_at')

        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Partners"

        # Define headers
        headers = [
            'Partner Name', 'Company Name', 'Person Name', 'Email', 'Phone',
            'Address', 'Country', 'Created At'
        ]

        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add data rows
        for row_num, partner in enumerate(partners, 2):
            data = [
                partner.partner_name,
                partner.company_name,
                partner.person_name,
                partner.email,
                partner.phone,
                partner.address,
                partner.country,
                partner.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ]

            for col_num, value in enumerate(data, 1):
                worksheet.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="partners_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        # Save workbook to response
        workbook.save(response)
        return response

    except Exception as e:
        messages.error(request, f'Error exporting partners: {str(e)}')
        return redirect('partners')